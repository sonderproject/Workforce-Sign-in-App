"""
Excel integration.

Responsibilities:
  * Back up the workbook before any write.
  * Resolve the correct intake sheet and Sign-In sheet for a reporting month.
  * Write the month's intake + visit data using the workbook's existing coding
    (``1`` / blank for categories, integer for House Size).
  * Never touch historical months.

Design note on duplicates
-------------------------
The database is the source of truth. For the *current* reporting month the app
owns its two sheets ("<Month> <Year>" and "<Month> <Year> Sign Ins") and
rebuilds them from the database on every sync. Rebuilding is idempotent and
therefore cannot create duplicate monthly-intake rows or duplicate visits, and
it leaves every other (historical) sheet completely untouched.
"""

from __future__ import annotations

import calendar
import os
import shutil
from datetime import datetime
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..mapping import (
    INTAKE_COLUMNS,
    SIGNIN_COLUMNS,
    TOTALS_ROW_LABEL,
    intake_cell_value,
    signin_cell_value,
)
from ..models import parse_dob

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
DATA_DIR = os.path.join(BASE_DIR, "data")
BACKUP_DIR = os.path.join(BASE_DIR, "backups")
DEFAULT_WORKBOOK = os.path.join(DATA_DIR, "reporting.xlsx")

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="2F5597")
_TOTALS_FONT = Font(bold=True)
_CENTER = Alignment(horizontal="center")
_DATE_FMT = "m/d/yyyy"


# ---------------------------------------------------------------------------
# Sheet naming
# ---------------------------------------------------------------------------

def month_sheet_name(year: int, month: int) -> str:
    return f"{calendar.month_name[month]} {year}"


def signin_sheet_name(year: int, month: int) -> str:
    return f"{calendar.month_name[month]} {year} Sign Ins"


# ---------------------------------------------------------------------------
# Backup
# ---------------------------------------------------------------------------

def backup_workbook(path: str) -> Optional[str]:
    """Copy ``path`` to backups/ with a timestamp. Returns the backup path."""
    if not os.path.exists(path):
        return None
    os.makedirs(BACKUP_DIR, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = os.path.splitext(os.path.basename(path))[0]
    dest = os.path.join(BACKUP_DIR, f"{base}_backup_{stamp}.xlsx")
    shutil.copy2(path, dest)
    return dest


# ---------------------------------------------------------------------------
# Data fetch from the database
# ---------------------------------------------------------------------------

def _fetch_month_intakes(conn, year, month):
    return conn.execute(
        """
        SELECT c.first_name, c.last_name, c.date_of_birth, c.uid,
               mi.veteran, mi.hispanic, mi.race, mi.house_size,
               mi.household_type, mi.population_category,
               mi.female_head, mi.disabled
        FROM monthly_intakes mi
        JOIN clients c ON c.id = mi.client_id
        WHERE mi.reporting_year = ? AND mi.reporting_month = ?
        ORDER BY c.last_norm, c.first_norm
        """,
        (year, month),
    ).fetchall()


def _fetch_month_visits(conn, year, month):
    return conn.execute(
        """
        SELECT c.first_name, c.last_name, c.date_of_birth,
               v.visit_date, v.visit_time, v.visitor_type, v.services
        FROM visits v
        JOIN clients c ON c.id = v.client_id
        WHERE v.reporting_year = ? AND v.reporting_month = ?
        ORDER BY v.visit_date, v.visit_time
        """,
        (year, month),
    ).fetchall()


# ---------------------------------------------------------------------------
# Cell writing helpers
# ---------------------------------------------------------------------------

def _write_headers(ws, columns):
    for idx, (header, _kind) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER


def _apply_value(cell, kind, value):
    if value is None:
        cell.value = None
        return
    if kind.startswith("date:"):
        d = parse_dob(value)
        if d is not None:
            cell.value = d
            cell.number_format = _DATE_FMT
        else:
            cell.value = value
    else:
        cell.value = value
        if kind.startswith(("flag:", "race:", "hh:", "pop:", "int:")):
            cell.alignment = _CENTER


def _autosize(ws, columns):
    for idx, (header, _kind) in enumerate(columns, start=1):
        letter = get_column_letter(idx)
        width = max(len(str(header)) + 2, 10)
        ws.column_dimensions[letter].width = min(width, 26)


# ---------------------------------------------------------------------------
# Sheet rebuild
# ---------------------------------------------------------------------------

def _replace_sheet(wb: Workbook, name: str):
    """Remove a sheet if present and return a fresh one with the same name."""
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(title=name)


def rebuild_intake_sheet(wb: Workbook, conn, year: int, month: int) -> int:
    """(Re)build the month's intake sheet from the DB. Returns row count."""
    ws = _replace_sheet(wb, month_sheet_name(year, month))
    _write_headers(ws, INTAKE_COLUMNS)

    rows = _fetch_month_intakes(conn, year, month)
    r = 2
    for row in rows:
        client = {
            "first_name": row["first_name"],
            "last_name": row["last_name"],
            "date_of_birth": row["date_of_birth"],
            "uid": row["uid"],
        }
        intake = dict(row)
        for c_idx, (_header, kind) in enumerate(INTAKE_COLUMNS, start=1):
            value = intake_cell_value(kind, intake, client)
            _apply_value(ws.cell(row=r, column=c_idx), kind, value)
        r += 1

    _write_totals_row(ws, INTAKE_COLUMNS, first_data_row=2, last_data_row=r - 1)
    _autosize(ws, INTAKE_COLUMNS)
    ws.freeze_panes = "A2"
    return len(rows)


def _write_totals_row(ws, columns, first_data_row, last_data_row):
    """Add a TOTALS row with SUM() over the numeric/category columns."""
    if last_data_row < first_data_row:
        return  # no data rows -> no totals row
    total_row = last_data_row + 1
    ws.cell(row=total_row, column=1, value=TOTALS_ROW_LABEL).font = _TOTALS_FONT
    for c_idx, (_header, kind) in enumerate(columns, start=1):
        if kind.startswith(("flag:", "race:", "hh:", "pop:", "int:")):
            letter = get_column_letter(c_idx)
            cell = ws.cell(
                row=total_row, column=c_idx,
                value=f"=SUM({letter}{first_data_row}:{letter}{last_data_row})",
            )
            cell.font = _TOTALS_FONT
            cell.alignment = _CENTER


def rebuild_signin_sheet(wb: Workbook, conn, year: int, month: int) -> int:
    """(Re)build the month's Sign-In sheet from the DB. Returns visit count."""
    ws = _replace_sheet(wb, signin_sheet_name(year, month))
    _write_headers(ws, SIGNIN_COLUMNS)

    rows = _fetch_month_visits(conn, year, month)
    r = 2
    for row in rows:
        client = {
            "first_name": row["first_name"],
            "last_name": row["last_name"],
            "date_of_birth": row["date_of_birth"],
        }
        visit = dict(row)
        for c_idx, (_header, kind) in enumerate(SIGNIN_COLUMNS, start=1):
            value = signin_cell_value(kind, visit, client)
            _apply_value(ws.cell(row=r, column=c_idx), kind, value)
        r += 1

    _autosize(ws, SIGNIN_COLUMNS)
    ws.freeze_panes = "A2"
    return len(rows)


# ---------------------------------------------------------------------------
# High-level sync operations
# ---------------------------------------------------------------------------

def sync_month_in_workbook(wb: Workbook, conn, year: int, month: int) -> dict:
    intake_count = rebuild_intake_sheet(wb, conn, year, month)
    visit_count = rebuild_signin_sheet(wb, conn, year, month)
    return {
        "intake_sheet": month_sheet_name(year, month),
        "signin_sheet": signin_sheet_name(year, month),
        "intake_rows": intake_count,
        "visit_rows": visit_count,
    }


def _load_or_create(path: str) -> Workbook:
    if os.path.exists(path):
        return load_workbook(path)
    wb = Workbook()
    # drop the default empty sheet; month sheets are created on demand
    default = wb.active
    default.title = "_placeholder"
    return wb


def _cleanup_placeholder(wb: Workbook):
    if "_placeholder" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["_placeholder"]


def sync_default_workbook(conn, year: int, month: int,
                          workbook_path: Optional[str] = None) -> dict:
    """Back up and update the app's working workbook for one month."""
    workbook_path = workbook_path or DEFAULT_WORKBOOK
    os.makedirs(os.path.dirname(workbook_path), exist_ok=True)
    backup = backup_workbook(workbook_path)
    wb = _load_or_create(workbook_path)
    result = sync_month_in_workbook(wb, conn, year, month)
    _cleanup_placeholder(wb)
    wb.save(workbook_path)
    result["backup"] = backup
    result["path"] = workbook_path
    return result


def apply_to_uploaded_workbook(conn, src_path: str, dest_path: str,
                               year: int, month: int) -> dict:
    """Drag-and-drop flow: take an uploaded workbook, write the month's data
    into the correct sheets, save to ``dest_path``. Historical sheets are
    preserved untouched. A backup of the uploaded file is taken first.
    """
    backup = backup_workbook(src_path)
    wb = load_workbook(src_path)
    result = sync_month_in_workbook(wb, conn, year, month)
    _cleanup_placeholder(wb)
    wb.save(dest_path)
    result["backup"] = backup
    result["path"] = dest_path
    return result
